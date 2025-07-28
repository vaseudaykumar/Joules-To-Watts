import pandas as pd
import time
import os
import FreeSimpleGUI as sg
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging

logging.basicConfig(filename='Transaction Exception Report.log', filemode='w', level=logging.DEBUG)
debug_it=0

def print_to_log(log_string):
    global logs
    log = str(time.strftime('%H:%M:%S'))+": "+str(log_string)
    logging.info(time.strftime('%H:%M:%S')+ ": "+log_string)
    sg.EasyPrint(log)

tab1_layout = [
    [sg.Text("Transaction File", size=(15,1), auto_size_text=False, justification='right'),\
     sg.InputText('Select the Transaction File', key='input_for_transaction_file'),sg.FileBrowse()],

    [sg.Button('Exception Report')]
]

layout=[
    [sg.TabGroup([[sg.Tab('Exception Report', tab1_layout),
                   ]])],
    [sg.Button('Cancel')]
]

window = sg.Window('Exception Report Automation Tool_v1.0.1').Layout(layout)


def exception_report_creation(transaction_file):
    try:

        print_to_log('Reading the Transaction File')
        df_transaction_file = pd.read_excel(transaction_file, sheet_name="Exceptions")

        def add_exception_status(txn):
            # Add a new column based on conditions in column 'Txn Amount'
            if pd.isna(txn):
                return "Missing"
            elif txn> 100000:
                return "High Value Alert"
            elif txn < 0:
                return "Negative Amount"
            elif txn == 0:
                return "Zero Transaction"
            else:
                return "Normal"

        # Adding Exception Status
        print_to_log('Adding Exception status')
        df_transaction_file['Exception Status'] = df_transaction_file['Transaction Amount'].apply(add_exception_status)

        print_to_log('Creating output file path')
        folder=os.path.dirname(transaction_file)
        output_filepath = os.path.join(folder,'Exception Report.xlsx')
        print_to_log(f'Output file path: {output_filepath}')

        # Write back to Excel
        df_transaction_file.to_excel(output_filepath, sheet_name='Exceptions', index=False)

        def create_table():

            print_to_log('Loading the excel for formatting')
            wb = load_workbook(output_filepath)
            sheet_name='Exceptions'
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                print_to_log(f'Sheet named "{sheet_name}" not found')
                return

            # Identifying the table range
            table_range=f'A1:{chr(64+df_transaction_file.shape[1])}{df_transaction_file.shape[0]+1}'

            #Create the table
            table = Table(displayName='TransactionTable', ref=table_range)

            #Add a default style with stripped rows
            style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style

            ws.add_table(table)

            # Aligning the cells & autofitting the cells
            print_to_log('Aligning the cells to center & autofitting the cells')
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if len(str(cell.value))>max_length:
                            max_length = len(cell.value)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            #Saving the output file
            print_to_log('Saving back the excel file')
            wb.save(output_filepath)

        print_to_log('Formatting the output file, this may take a while')
        create_table()

    except:
        print_to_log('Error reading Transaction file')

while True:
    button, values = window.Read()
    print(button, values)
    if button in (None, 'Cancel'):
        window.Close()
        break

    elif button == 'Exception Report':
        print_to_log('Please wait, while generating Exception Report')
        exception_report_creation(values['input_for_transaction_file'])
        print_to_log('Completed creating Exception Report')
        window['input_for_transaction_file'].update('Select the Transaction File')
window.Close()
